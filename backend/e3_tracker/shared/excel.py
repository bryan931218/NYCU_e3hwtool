import io
from datetime import datetime
from typing import Any, Dict, List, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .constants import TAIPEI_TZ


def _weekday_name(dt: datetime) -> str:
    names = ["一", "二", "三", "四", "五", "六", "日"]
    return names[dt.weekday()] if 0 <= dt.weekday() <= 6 else ""


def build_excel(
    assignments: List[Dict[str, Any]],
    *,
    output_path: str = "待繳作業.xlsx",
    return_bytes: bool = False,
) -> Union[str, io.BytesIO]:
    """將作業清單輸出成 Excel 檔案。

    Args:
        assignments: 作業清單
        output_path: 輸出檔案路徑
        return_bytes: 若為 True，回傳 BytesIO 而非寫入檔案

    Returns:
        若 return_bytes=False 則回傳檔名，否則回傳 BytesIO 物件
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "待繳作業"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4A5568")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["課程", "作業", "截止", "狀態"]
    ws.append(headers)
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    overdue_assignments = [a for a in assignments if a.get("overdue") and not a.get("completed")]
    future_assignments = [a for a in assignments if not a.get("overdue") and not a.get("completed")]

    overdue_fill = PatternFill("solid", fgColor="F56565")
    overdue_row_fill = PatternFill("solid", fgColor="FFF5F5")
    section_fill = PatternFill("solid", fgColor="E2E8F0")
    subtitle_fill = PatternFill("solid", fgColor="CBD5F5")
    zebra2_fill = PatternFill("solid", fgColor="F7FAFC")

    def write_title_row(text: str, fill: PatternFill, *, big: bool = False) -> None:
        safe = ("'" + text) if text.startswith("=") else text
        ws.append([safe, "", "", ""])
        row_idx = ws.max_row
        row_border = Border(top=Side(style="thick"), bottom=Side(style="thick")) if big else None
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            if big and row_border:
                cell.border = row_border
        target = ws.cell(row=row_idx, column=1)
        target.font = Font(bold=True, size=14 if big else 12)
        target.alignment = Alignment(horizontal="left")
        if big:
            ws.row_dimensions[row_idx].height = 22

    def write_rows(items: List[Dict[str, Any]], fill: PatternFill) -> None:
        for entry in items:
            ws.append(
                [
                    entry.get("course_title", ""),
                    entry.get("title", ""),
                    entry.get("due_at", "") or "無截止",
                    "作業連結",
                ]
            )
            row_idx = ws.max_row
            for col_idx in range(1, 5):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                if col_idx in (1, 2):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                elif col_idx == 3:
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            link_cell = ws.cell(row=row_idx, column=4)
            url = entry.get("url", "")
            if isinstance(url, str) and url.startswith("http"):
                try:
                    link_cell.hyperlink = url
                    link_cell.style = "Hyperlink"
                except Exception:
                    pass

    future_by_date: Dict[str, List[Dict[str, Any]]] = {}
    for item in future_assignments:
        due_ts = item.get("due_ts")
        if due_ts:
            key = datetime.fromtimestamp(due_ts, tz=TAIPEI_TZ).strftime("%Y-%m-%d")
        else:
            key = "無期限"
        future_by_date.setdefault(key, []).append(item)

    ordered_dates = sorted(future_by_date.keys(), key=lambda x: x)

    if overdue_assignments:
        write_title_row("===== 逾期未繳 =====", overdue_fill, big=True)
        write_rows(overdue_assignments, overdue_row_fill)

    if overdue_assignments and ordered_dates:
        ws.append(["", "", "", ""])

    if ordered_dates:
        write_title_row("===== 未繳交 =====", section_fill, big=True)
        for key in ordered_dates:
            items = sorted(future_by_date.get(key, []), key=lambda x: x.get("due_ts") or float("inf"))
            if not items:
                continue
            subtitle = f"-- {key}"
            try:
                dt_for_week = datetime.strptime(key, "%Y-%m-%d").replace(tzinfo=TAIPEI_TZ)
                subtitle += f" (週{_weekday_name(dt_for_week)}) --"
            except ValueError:
                subtitle += " --"
            write_title_row(subtitle, subtitle_fill)
            write_rows(items, zebra2_fill)

    for col_idx in range(1, 5):
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        base = 18 if col_idx in (1, 3) else 28 if col_idx == 2 else 14
        width = max(base, min(max_len + 2, 80))
        if col_idx in (1, 2):
            width = min(int(round(width * 1.2)), 100)
        ws.column_dimensions[chr(ord("A") + col_idx - 1)].width = width

    ws.freeze_panes = "A2"

    if ws.max_row <= 1:
        ws.append(["", "", "", ""])

    if return_bytes:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    target = output_path or "待繳作業.xlsx"
    try:
        wb.save(target)
        return target
    except Exception:
        fallback = f"{target}.tmp"
        wb.save(fallback)
        return fallback
